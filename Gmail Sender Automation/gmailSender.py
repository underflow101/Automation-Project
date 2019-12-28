# Gmail Sender
# 아래 라이브러리들은 파이썬 기본 내장 라이브러리이므로 별도의 설치가 필요 없습니다.

from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib

# Google의 Gmail 같은 경우 아래의 STMP 설정을 그대로 쓰면 됩니다. 포트 번호도 바꿀 필요 없습니다.
# 이는 Google에서 설정한 것이므로 gmail, 혹은 gsuite 기준 그대로 쓰면 됩니다.
# 단, 회사 메일이나 네이버, 다음 같은 경우 다르므로 SMTP 서버에 대해 알아보는 게 좋습니다.

# Gmail 같은 경우 해당 코드가 바로는 돌아가지 않습니다. 경고창에 뜨는 google의 URL로 들어가서
# "보안 수준이 낮은 앱 허용"을 "활성화"로 바꿔주셔야 합니다.
# "보안 수준이 낮은 앱 허용"의 기본 상태는 "비활성화"입니다.

SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 465
# 보내는 메일 계정
SMTP_USER = "보내는 사람 메일 계정"
SMTP_PASSWORD = "보내는 사람 메일 비밀번호"
# 만약 아래 메일 유효성 검사 함수에서 False가 나오면 메일을 보내지 않습니다.
def is_valid(addr):
    import re
    if re.match('(^[a-zA-Z-0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)', addr):
        return True
    else:
        return False
# 이메일 보내기 함수
def send_mail(addr, subj_layout, cont_layout, attachment=None):
    if not is_valid(addr):
        print("Wrong email: " + addr)
        return
    
    # 텍스트 파일
    msg = MIMEMultipart("alternative")
    # 첨부파일이 있는 경우 mixed로 multipart 생성
    if attachment:
        msg = MIMEMultipart('mixed')
    msg["From"] = SMTP_USER
    msg["To"] = addr
    msg["Subject"] = subj_layout
    contents = cont_layout
    text = MIMEText(_text = contents, _charset = "utf-8")
    msg.attach(text)

    if attachment:
        from email.mime.base import MIMEBase
        from email import encoders
        file_data = MIMEBase("application", "octect-stream")
        file_data.set_payload(open(attachment, "rb").read())
        encoders.encode_base64(file_data)
        import os
        filename = os.path.basename(attachment)
        file_data.add_header("Content-Disposition", 'attachment', filename=('UTF-8', '', filename))
        msg.attach(file_data)
    # smtp로 접속할 서버 정보를 가진 클래스변수 생성
    smtp = smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT)
    # 해당 서버로 로그인
    smtp.login(SMTP_USER, SMTP_PASSWORD)
    # 메일 발송
    smtp.sendmail(SMTP_USER, addr, msg.as_string())
    # 닫기
    smtp.close()

# 한 명에게 하나의 메일을 보낼 때
send_mail("받는 사람 주소", "메일 제목", "내용", "첨부파일 경로(생략 가능)")

# 한 명에게 여러 메일을 보낼 때
for i in range(100):
    send_mail("받는 사람 주소", "메일 제목", "내용", "첨부파일 경로(생략 가능)")

# 엑셀 파일에 정리된 명단으로 한꺼번에 보낼 때
# 아래 openpyxl 라이브러리는 외부 라이브러리이므로 pip3를 통해 설치 후 사용하시기 바랍니다.
from openpyxl import load_workbook
wb = load_workbook('ex.xlsx')
ws = wb.active
for row in ws.iter_rows():
    addr = row[0].value
    subj_layout = row[1].value
    cont_layout = row[2].value
    attachment = row[3].value
    send_mail(addr, subj_layout, cont_layout, attachment)