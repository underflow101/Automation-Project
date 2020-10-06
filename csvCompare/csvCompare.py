import openpyxl
from openpyxl import Workbook
from collections import deque

origin_dir = '/home/bearpaek/data/ksOlly/origin.xlsx'
new_dir = '/home/bearpaek/data/ksOlly/new.xlsx'
brand_dir = '/home/bearpaek/data/ksOlly/brand.xlsx'

visited = []

wb = Workbook()
sheet1 = wb.active

origin_xl = openpyxl.load_workbook(origin_dir)
new_xl = openpyxl.load_workbook(new_dir)

origin_sheet = origin_xl.worksheets[0]
new_sheet = new_xl.worksheets[0]

i = 1
j = 1

for origin_row in origin_sheet.iter_rows(min_row=1501, max_row=1918, min_col=5, max_col=5):
    for origin_cell in origin_row:
        j = 1
        for new_row in new_sheet.iter_rows(min_row=1, max_row=1299, min_col=4, max_col=4):
            for new_cell in new_row:
                if origin_cell.value == new_cell.value:
                        k = 0
                        tmplst = [''] * 9
                        tmpstr = ''
                        for tmp in new_sheet.iter_rows(min_row=j, max_row=j, min_col=1, max_col=25):
                            for cell in tmp:
                                k += 1
                                if k == 16:
                                    if cell.value == None:
                                        continue
                                    tmplst[0] = str(cell.value)
                                elif k == 24:
                                    if cell.value == None:
                                        continue
                                    tmplst[1] = str(cell.value)
                                elif k == 4:
                                    if cell.value == None:
                                        continue
                                    tmplst[2] = str(cell.value)
                                elif k == 17:
                                    tmpstr += str(cell.value)
                                    tmpstr += ' '
                                elif k == 18:
                                    if cell.value == None:
                                        tmplst[4] = tmpstr
                                        continue
                                    tmpstr += str(cell.value)
                                    tmplst[4] = tmpstr
                                elif k == 19:
                                    if cell.value == None:
                                        continue
                                    tmplst[5] = str(cell.value)
                                elif k == 20:
                                    if cell.value == None:
                                        continue
                                    tmplst[6] = str(cell.value)
                                elif k == 21:
                                    if cell.value == None:
                                        continue
                                    tmplst[7] = str(cell.value)
                                elif k == 23:
                                    if cell.value == None:
                                        continue
                                    tmplst[8] = str(cell.value)
                        tmplst = deque(tmplst)
                        for ii in range(len(tmplst)):
                            print(tmplst.popleft(), end=';')
                        print('')
                        visited.append(origin_cell.value)
            j += 1
        